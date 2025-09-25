# Excel import/export utilities
import pandas as pd
from sqlalchemy.orm import Session
from sqlalchemy import text
from typing import Optional, Dict, Any
import io
import os
from datetime import datetime
import logging

from .db_models import User, Company, Contact, Project, VOC, AuditLog

logger = logging.getLogger(__name__)

# Export 디렉토리 경로 (프로젝트 루트 기준)
EXPORT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "exports")

def ensure_export_directory():
    """Export 디렉토리가 존재하는지 확인하고, 없으면 생성"""
    if not os.path.exists(EXPORT_DIR):
        os.makedirs(EXPORT_DIR)
        logger.info(f"Created export directory: {EXPORT_DIR}")
    return EXPORT_DIR

def export_voc_to_excel(db: Session, filename: Optional[str] = None) -> str:
    """
    VOC 테이블만 엑셀로 내보내기
    assignee_user_id, company_id, contact_id, project_id, deleted_at 컬럼 제외
    
    Args:
        db: 데이터베이스 세션
        filename: 파일명 (선택사항, 기본값은 export_voc_25XXXX.xlsx)
    
    Returns:
        str: 생성된 엑셀 파일의 전체 경로
    """
    try:
        # Export 디렉토리 확인 및 생성
        export_dir = ensure_export_directory()

        # 파일명 생성 (25XXXX 형식)
        if filename is None:
            timestamp = datetime.now().strftime("%y%m%d_%H%M")
            filename = f"export_voc_{timestamp}.xlsx"

        # 전체 파일 경로 생성
        filepath = os.path.join(export_dir, filename)

        # ExcelWriter 생성
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # VOC 테이블만 내보내기 (특정 컬럼 제외)
            _export_voc_table_to_sheet(db, "VOCs", writer)

        logger.info(f"VOC Excel export completed: {filepath}")
        return filepath

    except Exception as e:
        logger.error(f"VOC Excel export failed: {str(e)}")
        raise

def export_full_tables_to_excel(db: Session, filename: Optional[str] = None) -> str:
    """
    Users 제외한 모든 테이블을 엑셀로 내보내기
    연동된 ID 컬럼들 제외 (assignee_user_id, company_id, contact_id, project_id 등)
    
    Args:
        db: 데이터베이스 세션
        filename: 파일명 (선택사항, 기본값은 export_full_25XXXX.xlsx)
    
    Returns:
        str: 생성된 엑셀 파일의 전체 경로
    """
    try:
        # Export 디렉토리 확인 및 생성
        export_dir = ensure_export_directory()

        # 파일명 생성 (25XXXX 형식)
        if filename is None:
            timestamp = datetime.now().strftime("%y%m%d_%H%M")
            filename = f"export_full_{timestamp}.xlsx"

        # 전체 파일 경로 생성
        filepath = os.path.join(export_dir, filename)

        # ExcelWriter 생성
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Users 제외한 테이블들 내보내기 (연동 ID 제외)
            _export_table_to_sheet_exclude_ids(db, Company, "Companies", writer)
            _export_table_to_sheet_exclude_ids(db, Contact, "Contacts", writer)
            _export_table_to_sheet_exclude_ids(db, Project, "Projects", writer)
            _export_table_to_sheet_exclude_ids(db, VOC, "VOCs", writer)
            _export_table_to_sheet_exclude_ids(db, AuditLog, "AuditLogs", writer)

        logger.info(f"Full Excel export completed: {filepath}")
        return filepath

    except Exception as e:
        logger.error(f"Full Excel export failed: {str(e)}")
        raise

def export_biz_template_to_excel(db: Session, filename: Optional[str] = None) -> str:
    """
    VOC와 Projects 2개 시트가 있는 비즈니스 템플릿 엑셀 파일 생성
    나중에 input 템플릿으로 사용할 예정
    - VOC 시트: company_name 추가, create/update 숨기기, 간격 조정
    - Projects 시트: company_name 추가, create/update 숨기기, 간격 조정
    
    Args:
        db: 데이터베이스 세션
        filename: 파일명 (선택사항, 기본값은 export_biz_25XXXX.xlsx)
    
    Returns:
        str: 생성된 엑셀 파일의 전체 경로
    """
    try:
        # Export 디렉토리 확인 및 생성
        export_dir = ensure_export_directory()

        # 파일명 생성 (25XXXX 형식)
        if filename is None:
            timestamp = datetime.now().strftime("%y%m%d_%H%M")
            filename = f"export_biz_{timestamp}.xlsx"

        # 전체 파일 경로 생성
        filepath = os.path.join(export_dir, filename)

        # ExcelWriter 생성
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            try:
                # VOC 시트 (템플릿용 포맷팅) - 데이터베이스 연결 시도
                _export_voc_template_to_sheet(db, "VOCs", writer)
            except Exception as e:
                logger.warning(f"Database connection failed, creating empty VOC template: {e}")
                # 데이터베이스 연결 실패 시 빈 템플릿 생성
                _create_empty_voc_template_sheet(writer, "VOCs")
            
            try:
                # Projects 시트 (템플릿용 포맷팅) - 데이터베이스 연결 시도
                _export_project_template_to_sheet(db, "Projects", writer)
            except Exception as e:
                logger.warning(f"Database connection failed, creating empty Project template: {e}")
                # 데이터베이스 연결 실패 시 빈 템플릿 생성
                _create_empty_project_template_sheet(writer, "Projects")

        logger.info(f"Business template Excel export completed: {filepath}")
        return filepath

    except Exception as e:
        logger.error(f"Business template Excel export failed: {str(e)}")
        raise

def export_all_tables_to_excel(db: Session, filename: Optional[str] = None) -> str:
    """
    모든 테이블의 데이터를 엑셀로 내보내기 (기존 기능 유지)
    각 테이블은 별도의 탭으로 생성됩니다.

    Args:
        db: 데이터베이스 세션
        filename: 파일명 (선택사항, 기본값은 export_all_25XXXX.xlsx)

    Returns:
        str: 생성된 엑셀 파일의 전체 경로
    """
    try:
        # Export 디렉토리 확인 및 생성
        export_dir = ensure_export_directory()

        # 파일명 생성 (25XXXX 형식)
        if filename is None:
            timestamp = datetime.now().strftime("%y%m%d_%H%M")
            filename = f"export_all_{timestamp}.xlsx"

        # 전체 파일 경로 생성
        filepath = os.path.join(export_dir, filename)

        # ExcelWriter 생성
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # 각 테이블별로 데이터 추출 및 엑셀 시트 생성
            _export_table_to_sheet(db, User, "Users", writer)
            _export_table_to_sheet(db, Company, "Companies", writer)
            _export_table_to_sheet(db, Contact, "Contacts", writer)
            _export_table_to_sheet(db, Project, "Projects", writer)
            _export_table_to_sheet(db, VOC, "VOCs", writer)
            _export_table_to_sheet(db, AuditLog, "AuditLogs", writer)

        logger.info(f"Excel export completed: {filepath}")
        return filepath

    except Exception as e:
        logger.error(f"Excel export failed: {str(e)}")
        raise

def _export_table_to_sheet(db: Session, model_class, sheet_name: str, writer: pd.ExcelWriter):
    """
    특정 테이블의 데이터를 엑셀 시트로 내보내기
    
    Args:
        db: 데이터베이스 세션
        model_class: SQLAlchemy 모델 클래스
        sheet_name: 엑셀 시트명
        writer: pandas ExcelWriter 객체
    """
    try:
        # 테이블의 모든 데이터 조회
        query = db.query(model_class)
        results = query.all()
        
        if not results:
            # 데이터가 없는 경우 빈 DataFrame 생성
            df = pd.DataFrame(columns=[column.name for column in model_class.__table__.columns])
        else:
            # 결과를 딕셔너리 리스트로 변환
            data = []
            for result in results:
                row_dict = {}
                for column in model_class.__table__.columns:
                    value = getattr(result, column.name)
                    # datetime 객체를 문자열로 변환
                    if hasattr(value, 'strftime'):
                        value = value.strftime('%Y-%m-%d %H:%M:%S') if hasattr(value, 'hour') else value.strftime('%Y-%m-%d')
                    row_dict[column.name] = value
                data.append(row_dict)
            
            df = pd.DataFrame(data)
        
        # 엑셀 시트로 저장
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 컬럼 너비 자동 조정
        _adjust_column_widths(writer, sheet_name, df)
        
        logger.info(f"Exported {len(results)} records from {model_class.__tablename__} to sheet '{sheet_name}'")
        
    except Exception as e:
        logger.error(f"Failed to export {model_class.__tablename__} to sheet '{sheet_name}': {str(e)}")
        # 에러가 발생해도 빈 시트라도 생성
        empty_df = pd.DataFrame(columns=[column.name for column in model_class.__table__.columns])
        empty_df.to_excel(writer, sheet_name=sheet_name, index=False)

def _export_voc_table_to_sheet(db: Session, sheet_name: str, writer: pd.ExcelWriter):
    """
    VOC 테이블을 엑셀 시트로 내보내기 (특정 컬럼 제외, company_name 추가)
    assignee_user_id, company_id, contact_id, project_id, deleted_at 컬럼 제외
    company_name을 date 컬럼 오른쪽에 추가
    
    Args:
        db: 데이터베이스 세션
        sheet_name: 엑셀 시트명
        writer: pandas ExcelWriter 객체
    """
    try:
        # VOC 테이블의 모든 데이터 조회 (Company와 조인)
        query = db.query(VOC).join(Company, VOC.company_id == Company.id)
        results = query.all()
        
        # 제외할 컬럼들
        exclude_columns = {'assignee_user_id', 'company_id', 'contact_id', 'project_id', 'deleted_at'}
        
        if not results:
            # 데이터가 없는 경우 빈 DataFrame 생성 (제외 컬럼 제외, company_name 추가)
            all_columns = [column.name for column in VOC.__table__.columns]
            filtered_columns = [col for col in all_columns if col not in exclude_columns]
            # company_name을 date 오른쪽에 추가
            if 'date' in filtered_columns:
                date_index = filtered_columns.index('date')
                filtered_columns.insert(date_index + 1, 'company_name')
            else:
                filtered_columns.append('company_name')
            df = pd.DataFrame(columns=filtered_columns)
        else:
            # 결과를 딕셔너리 리스트로 변환 (제외 컬럼 제외, company_name 추가)
            data = []
            for result in results:
                row_dict = {}
                for column in VOC.__table__.columns:
                    if column.name not in exclude_columns:
                        value = getattr(result, column.name)
                        # datetime 객체를 문자열로 변환
                        if hasattr(value, 'strftime'):
                            value = value.strftime('%Y-%m-%d %H:%M:%S') if hasattr(value, 'hour') else value.strftime('%Y-%m-%d')
                        row_dict[column.name] = value
                
                # company_name 추가 (date 오른쪽에)
                if hasattr(result, 'company') and result.company:
                    row_dict['company_name'] = result.company.name
                else:
                    row_dict['company_name'] = None
                
                data.append(row_dict)
            
            df = pd.DataFrame(data)
            
            # 컬럼 순서 조정: company_name을 date 오른쪽으로 이동
            if 'date' in df.columns and 'company_name' in df.columns:
                cols = df.columns.tolist()
                # date와 company_name의 위치 찾기
                date_idx = cols.index('date')
                company_name_idx = cols.index('company_name')
                
                # company_name을 date 오른쪽으로 이동
                cols.pop(company_name_idx)
                cols.insert(date_idx + 1, 'company_name')
                df = df[cols]
        
        # 엑셀 시트로 저장
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 컬럼 너비 자동 조정
        _adjust_column_widths(writer, sheet_name, df)
        
        logger.info(f"Exported {len(results)} VOC records to sheet '{sheet_name}' (excluding {len(exclude_columns)} columns, added company_name)")
        
    except Exception as e:
        logger.error(f"Failed to export VOC table to sheet '{sheet_name}': {str(e)}")
        # 에러가 발생해도 빈 시트라도 생성
        all_columns = [column.name for column in VOC.__table__.columns]
        exclude_columns = {'assignee_user_id', 'company_id', 'contact_id', 'project_id', 'deleted_at'}
        filtered_columns = [col for col in all_columns if col not in exclude_columns]
        # company_name 추가
        if 'date' in filtered_columns:
            date_index = filtered_columns.index('date')
            filtered_columns.insert(date_index + 1, 'company_name')
        else:
            filtered_columns.append('company_name')
        empty_df = pd.DataFrame(columns=filtered_columns)
        empty_df.to_excel(writer, sheet_name=sheet_name, index=False)

def _export_voc_template_to_sheet(db: Session, sheet_name: str, writer: pd.ExcelWriter):
    """
    VOC 테이블을 템플릿용 엑셀 시트로 내보내기
    - company_name 추가 (date 오른쪽)
    - created_at, updated_at 숨기기
    - 템플릿용 간격 조정 및 포맷팅
    
    Args:
        db: 데이터베이스 세션
        sheet_name: 엑셀 시트명
        writer: pandas ExcelWriter 객체
    """
    try:
        # VOC 테이블의 모든 데이터 조회 (Company와 조인)
        query = db.query(VOC).join(Company, VOC.company_id == Company.id)
        results = query.all()
        
        # 제외할 컬럼들 (템플릿용)
        exclude_columns = {'assignee_user_id', 'company_id', 'contact_id', 'project_id', 'deleted_at', 'created_at', 'updated_at'}
        
        if not results:
            # 데이터가 없는 경우 빈 DataFrame 생성 (제외 컬럼 제외, company_name 추가)
            all_columns = [column.name for column in VOC.__table__.columns]
            filtered_columns = [col for col in all_columns if col not in exclude_columns]
            # company_name을 date 오른쪽에 추가
            if 'date' in filtered_columns:
                date_index = filtered_columns.index('date')
                filtered_columns.insert(date_index + 1, 'company_name')
            else:
                filtered_columns.append('company_name')
            df = pd.DataFrame(columns=filtered_columns)
        else:
            # 결과를 딕셔너리 리스트로 변환 (제외 컬럼 제외, company_name 추가)
            data = []
            for result in results:
                row_dict = {}
                for column in VOC.__table__.columns:
                    if column.name not in exclude_columns:
                        value = getattr(result, column.name)
                        # datetime 객체를 문자열로 변환
                        if hasattr(value, 'strftime'):
                            value = value.strftime('%Y-%m-%d %H:%M:%S') if hasattr(value, 'hour') else value.strftime('%Y-%m-%d')
                        row_dict[column.name] = value
                
                # company_name 추가 (date 오른쪽에)
                if hasattr(result, 'company') and result.company:
                    row_dict['company_name'] = result.company.name
                else:
                    row_dict['company_name'] = None
                
                data.append(row_dict)
            
            df = pd.DataFrame(data)
            
            # 컬럼 순서 조정: company_name을 date 오른쪽으로 이동
            if 'date' in df.columns and 'company_name' in df.columns:
                cols = df.columns.tolist()
                # date와 company_name의 위치 찾기
                date_idx = cols.index('date')
                company_name_idx = cols.index('company_name')
                
                # company_name을 date 오른쪽으로 이동
                cols.pop(company_name_idx)
                cols.insert(date_idx + 1, 'company_name')
                df = df[cols]
        
        # 엑셀 시트로 저장
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 템플릿용 컬럼 너비 및 포맷팅 조정
        _adjust_template_column_widths(writer, sheet_name, df, 'voc')
        
        logger.info(f"Exported {len(results)} VOC records to template sheet '{sheet_name}' (excluding {len(exclude_columns)} columns, added company_name)")
        
    except Exception as e:
        logger.error(f"Failed to export VOC template to sheet '{sheet_name}': {str(e)}")
        # 에러가 발생해도 빈 시트라도 생성
        all_columns = [column.name for column in VOC.__table__.columns]
        exclude_columns = {'assignee_user_id', 'company_id', 'contact_id', 'project_id', 'deleted_at', 'created_at', 'updated_at'}
        filtered_columns = [col for col in all_columns if col not in exclude_columns]
        # company_name 추가
        if 'date' in filtered_columns:
            date_index = filtered_columns.index('date')
            filtered_columns.insert(date_index + 1, 'company_name')
        else:
            filtered_columns.append('company_name')
        empty_df = pd.DataFrame(columns=filtered_columns)
        empty_df.to_excel(writer, sheet_name=sheet_name, index=False)

def _export_project_template_to_sheet(db: Session, sheet_name: str, writer: pd.ExcelWriter):
    """
    Project 테이블을 템플릿용 엑셀 시트로 내보내기
    - company_name 추가 (name 오른쪽)
    - created_at, updated_at 숨기기
    - 템플릿용 간격 조정 및 포맷팅
    
    Args:
        db: 데이터베이스 세션
        sheet_name: 엑셀 시트명
        writer: pandas ExcelWriter 객체
    """
    try:
        # Project 테이블의 모든 데이터 조회 (Company와 조인)
        query = db.query(Project).join(Company, Project.company_id == Company.id)
        results = query.all()
        
        # 제외할 컬럼들 (템플릿용)
        exclude_columns = {'company_id', 'created_at', 'updated_at'}
        
        if not results:
            # 데이터가 없는 경우 빈 DataFrame 생성 (제외 컬럼 제외, company_name 추가)
            all_columns = [column.name for column in Project.__table__.columns]
            filtered_columns = [col for col in all_columns if col not in exclude_columns]
            # company_name을 name 오른쪽에 추가
            if 'name' in filtered_columns:
                name_index = filtered_columns.index('name')
                filtered_columns.insert(name_index + 1, 'company_name')
            else:
                filtered_columns.append('company_name')
            df = pd.DataFrame(columns=filtered_columns)
        else:
            # 결과를 딕셔너리 리스트로 변환 (제외 컬럼 제외, company_name 추가)
            data = []
            for result in results:
                row_dict = {}
                for column in Project.__table__.columns:
                    if column.name not in exclude_columns:
                        value = getattr(result, column.name)
                        # datetime 객체를 문자열로 변환
                        if hasattr(value, 'strftime'):
                            value = value.strftime('%Y-%m-%d %H:%M:%S') if hasattr(value, 'hour') else value.strftime('%Y-%m-%d')
                        row_dict[column.name] = value
                
                # company_name 추가 (name 오른쪽에)
                if hasattr(result, 'company') and result.company:
                    row_dict['company_name'] = result.company.name
                else:
                    row_dict['company_name'] = None
                
                data.append(row_dict)
            
            df = pd.DataFrame(data)
            
            # 컬럼 순서 조정: company_name을 name 오른쪽으로 이동
            if 'name' in df.columns and 'company_name' in df.columns:
                cols = df.columns.tolist()
                # name과 company_name의 위치 찾기
                name_idx = cols.index('name')
                company_name_idx = cols.index('company_name')
                
                # company_name을 name 오른쪽으로 이동
                cols.pop(company_name_idx)
                cols.insert(name_idx + 1, 'company_name')
                df = df[cols]
        
        # 엑셀 시트로 저장
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 템플릿용 컬럼 너비 및 포맷팅 조정
        _adjust_template_column_widths(writer, sheet_name, df, 'project')
        
        logger.info(f"Exported {len(results)} Project records to template sheet '{sheet_name}' (excluding {len(exclude_columns)} columns, added company_name)")
        
    except Exception as e:
        logger.error(f"Failed to export Project template to sheet '{sheet_name}': {str(e)}")
        # 에러가 발생해도 빈 시트라도 생성
        all_columns = [column.name for column in Project.__table__.columns]
        exclude_columns = {'company_id', 'created_at', 'updated_at'}
        filtered_columns = [col for col in all_columns if col not in exclude_columns]
        # company_name 추가
        if 'name' in filtered_columns:
            name_index = filtered_columns.index('name')
            filtered_columns.insert(name_index + 1, 'company_name')
        else:
            filtered_columns.append('company_name')
        empty_df = pd.DataFrame(columns=filtered_columns)
        empty_df.to_excel(writer, sheet_name=sheet_name, index=False)

def _create_empty_voc_template_sheet(writer: pd.ExcelWriter, sheet_name: str):
    """
    빈 VOC 템플릿 시트 생성 (데이터베이스 연결 없이)
    - company_name 추가, create/update 숨기기, 간격 조정
    
    Args:
        writer: pandas ExcelWriter 객체
        sheet_name: 엑셀 시트명
    """
    try:
        # VOC 템플릿용 컬럼 정의 (제외 컬럼 제외, company_name 추가)
        voc_columns = [
            'id', 'date', 'company_name', 'content', 'action_item', 
            'status', 'priority', 'due_date', 'ai_summary'
        ]
        
        # 빈 DataFrame 생성
        df = pd.DataFrame(columns=voc_columns)
        
        # 엑셀 시트로 저장
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 템플릿용 컬럼 너비 및 포맷팅 조정
        _adjust_template_column_widths(writer, sheet_name, df, 'voc')
        
        logger.info(f"Created empty VOC template sheet '{sheet_name}' with formatting")
        
    except Exception as e:
        logger.error(f"Failed to create empty VOC template sheet '{sheet_name}': {str(e)}")
        # 에러가 발생해도 기본 시트라도 생성
        voc_columns = ['id', 'date', 'company_name', 'content', 'action_item', 'status', 'priority']
        empty_df = pd.DataFrame(columns=voc_columns)
        empty_df.to_excel(writer, sheet_name=sheet_name, index=False)

def _create_empty_project_template_sheet(writer: pd.ExcelWriter, sheet_name: str):
    """
    빈 Project 템플릿 시트 생성 (데이터베이스 연결 없이)
    - company_name 추가, create/update 숨기기, 간격 조정
    
    Args:
        writer: pandas ExcelWriter 객체
        sheet_name: 엑셀 시트명
    """
    try:
        # Project 템플릿용 컬럼 정의 (제외 컬럼 제외, company_name 추가)
        project_columns = [
            'id', 'name', 'company_name', 'field', 'target_app', 'ai_model',
            'perf', 'power', 'size', 'price', 'requirements', 'competitors', 
            'result', 'root_cause'
        ]
        
        # 빈 DataFrame 생성
        df = pd.DataFrame(columns=project_columns)
        
        # 엑셀 시트로 저장
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 템플릿용 컬럼 너비 및 포맷팅 조정
        _adjust_template_column_widths(writer, sheet_name, df, 'project')
        
        logger.info(f"Created empty Project template sheet '{sheet_name}' with formatting")
        
    except Exception as e:
        logger.error(f"Failed to create empty Project template sheet '{sheet_name}': {str(e)}")
        # 에러가 발생해도 기본 시트라도 생성
        project_columns = ['id', 'name', 'company_name', 'field', 'target_app', 'ai_model', 'requirements']
        empty_df = pd.DataFrame(columns=project_columns)
        empty_df.to_excel(writer, sheet_name=sheet_name, index=False)

def _export_table_to_sheet_exclude_ids(db: Session, model_class, sheet_name: str, writer: pd.ExcelWriter):
    """
    특정 테이블의 데이터를 엑셀 시트로 내보내기 (연동 ID 컬럼들 제외)
    
    Args:
        db: 데이터베이스 세션
        model_class: SQLAlchemy 모델 클래스
        sheet_name: 엑셀 시트명
        writer: pandas ExcelWriter 객체
    """
    try:
        # 테이블의 모든 데이터 조회
        query = db.query(model_class)
        results = query.all()
        
        # 제외할 ID 컬럼들 (테이블별로 다름)
        exclude_columns = set()
        
        if model_class == Company:
            # Company는 ID 컬럼 제외 없음
            pass
        elif model_class == Contact:
            exclude_columns = {'company_id'}
        elif model_class == Project:
            exclude_columns = {'company_id'}
        elif model_class == VOC:
            exclude_columns = {'assignee_user_id', 'company_id', 'contact_id', 'project_id', 'deleted_at'}
        elif model_class == AuditLog:
            exclude_columns = {'actor_user_id'}
        
        if not results:
            # 데이터가 없는 경우 빈 DataFrame 생성 (제외 컬럼 제외)
            all_columns = [column.name for column in model_class.__table__.columns]
            filtered_columns = [col for col in all_columns if col not in exclude_columns]
            df = pd.DataFrame(columns=filtered_columns)
        else:
            # 결과를 딕셔너리 리스트로 변환 (제외 컬럼 제외)
            data = []
            for result in results:
                row_dict = {}
                for column in model_class.__table__.columns:
                    if column.name not in exclude_columns:
                        value = getattr(result, column.name)
                        # datetime 객체를 문자열로 변환
                        if hasattr(value, 'strftime'):
                            value = value.strftime('%Y-%m-%d %H:%M:%S') if hasattr(value, 'hour') else value.strftime('%Y-%m-%d')
                        row_dict[column.name] = value
                data.append(row_dict)
            
            df = pd.DataFrame(data)
        
        # 엑셀 시트로 저장
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 컬럼 너비 자동 조정
        _adjust_column_widths(writer, sheet_name, df)
        
        excluded_info = f" (excluding {len(exclude_columns)} ID columns)" if exclude_columns else ""
        logger.info(f"Exported {len(results)} records from {model_class.__tablename__} to sheet '{sheet_name}'{excluded_info}")
        
    except Exception as e:
        logger.error(f"Failed to export {model_class.__tablename__} to sheet '{sheet_name}': {str(e)}")
        # 에러가 발생해도 빈 시트라도 생성
        all_columns = [column.name for column in model_class.__table__.columns]
        exclude_columns = set()
        
        if model_class == Contact:
            exclude_columns = {'company_id'}
        elif model_class == Project:
            exclude_columns = {'company_id'}
        elif model_class == VOC:
            exclude_columns = {'assignee_user_id', 'company_id', 'contact_id', 'project_id', 'deleted_at'}
        elif model_class == AuditLog:
            exclude_columns = {'actor_user_id'}
        
        filtered_columns = [col for col in all_columns if col not in exclude_columns]
        empty_df = pd.DataFrame(columns=filtered_columns)
        empty_df.to_excel(writer, sheet_name=sheet_name, index=False)

def _adjust_column_widths(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame):
    """
    엑셀 컬럼 너비 자동 조정
    
    Args:
        writer: pandas ExcelWriter 객체
        sheet_name: 엑셀 시트명
        df: DataFrame
    """
    try:
        from openpyxl.utils import get_column_letter
        
        # 워크시트 가져오기
        worksheet = writer.sheets[sheet_name]
        
        # 컬럼별 너비 설정
        for column in df.columns:
            col_letter = get_column_letter(df.columns.get_loc(column) + 1)
            
            # 컬럼명 길이 계산
            col_name_length = len(str(column))
            
            # 데이터 최대 길이 계산
            max_data_length = 0
            if not df.empty:
                for value in df[column]:
                    if value is not None:
                        max_data_length = max(max_data_length, len(str(value)))
            
            # 최소 너비 설정
            min_width = max(col_name_length, 10)
            
            # 특별한 컬럼들은 더 넓게 설정
            if column in ['content', 'action_item', 'requirements', 'competitors', 'result', 'root_cause', 'note']:
                # 텍스트 컬럼은 최소 30, 최대 50
                width = min(max(min_width, max_data_length, 30), 50)
            elif column in ['ai_summary']:
                # AI 요약은 최소 25, 최대 40
                width = min(max(min_width, max_data_length, 25), 40)
            elif column in ['name', 'title', 'email', 'phone', 'domain', 'revenue', 'nation', 'field', 'target_app', 'ai_model', 'perf', 'power', 'size', 'price', 'company_name']:
                # 일반 텍스트 컬럼은 최소 15, 최대 25
                width = min(max(min_width, max_data_length, 15), 25)
            elif column in ['status', 'priority', 'action', 'table_name']:
                # 상태/타입 컬럼은 최소 12, 최대 18
                width = min(max(min_width, max_data_length, 12), 18)
            elif column in ['date', 'due_date']:
                # 날짜 컬럼은 딱 맞게 조정 (YYYY-MM-DD 형식)
                width = 12
            elif column in ['created_at', 'updated_at']:
                # 타임스탬프 컬럼은 숨김 처리
                width = 12
            elif column in ['id', 'employee']:
                # 숫자 컬럼은 최소 8, 최대 12
                width = min(max(min_width, max_data_length, 8), 12)
            else:
                # 기본 너비
                width = min(max(min_width, max_data_length, 12), 20)
            
            # 컬럼 너비 설정
            worksheet.column_dimensions[col_letter].width = width
            
            # created_at, updated_at 컬럼 숨기기
            if column in ['created_at', 'updated_at']:
                worksheet.column_dimensions[col_letter].hidden = True
            
            # content, action_item 컬럼에 자동 줄바꿈 설정
            if column in ['content', 'action_item']:
                for row in range(2, worksheet.max_row + 1):  # 헤더 제외하고 데이터 행부터
                    cell = worksheet[f"{col_letter}{row}"]
                    from openpyxl.styles import Alignment
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            
        logger.info(f"Adjusted column widths for sheet '{sheet_name}'")
        
    except Exception as e:
        logger.warning(f"Failed to adjust column widths for sheet '{sheet_name}': {str(e)}")

def _adjust_template_column_widths(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame, sheet_type: str):
    """
    템플릿용 엑셀 컬럼 너비 자동 조정 (VOC/Project 전용)
    - created_at, updated_at 컬럼 숨기기
    - 템플릿용 최적화된 너비 설정
    - 자동 줄바꿈 설정
    
    Args:
        writer: pandas ExcelWriter 객체
        sheet_name: 엑셀 시트명
        df: DataFrame
        sheet_type: 시트 타입 ('voc' 또는 'project')
    """
    try:
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment, Font, PatternFill
        
        # 워크시트 가져오기
        worksheet = writer.sheets[sheet_name]
        
        # 헤더 스타일 설정
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # 컬럼별 너비 설정
        for column in df.columns:
            col_letter = get_column_letter(df.columns.get_loc(column) + 1)
            
            # 컬럼명 길이 계산
            col_name_length = len(str(column))
            
            # 데이터 최대 길이 계산
            max_data_length = 0
            if not df.empty:
                for value in df[column]:
                    if value is not None:
                        max_data_length = max(max_data_length, len(str(value)))
            
            # 최소 너비 설정
            min_width = max(col_name_length, 10)
            
            # 템플릿용 특별한 컬럼 너비 설정
            if sheet_type == 'voc':
                if column in ['content', 'action_item']:
                    # VOC 본문/액션아이템은 넓게 (최소 40, 최대 60)
                    width = min(max(min_width, max_data_length, 40), 60)
                elif column in ['ai_summary']:
                    # AI 요약은 중간 크기 (최소 30, 최대 45)
                    width = min(max(min_width, max_data_length, 30), 45)
                elif column in ['company_name']:
                    # 회사명은 중간 크기 (최소 20, 최대 30)
                    width = min(max(min_width, max_data_length, 20), 30)
                elif column in ['date', 'due_date']:
                    # 날짜 컬럼은 딱 맞게
                    width = 12
                elif column in ['status', 'priority']:
                    # 상태/우선순위는 작게
                    width = min(max(min_width, max_data_length, 10), 15)
                elif column in ['id']:
                    # ID는 작게
                    width = 8
                else:
                    # 기본 너비
                    width = min(max(min_width, max_data_length, 15), 25)
            
            elif sheet_type == 'project':
                if column in ['requirements', 'competitors', 'result', 'root_cause']:
                    # 프로젝트 상세 정보는 넓게 (최소 35, 최대 50)
                    width = min(max(min_width, max_data_length, 35), 50)
                elif column in ['name', 'company_name']:
                    # 프로젝트명/회사명은 중간 크기 (최소 20, 최대 30)
                    width = min(max(min_width, max_data_length, 20), 30)
                elif column in ['field', 'target_app', 'ai_model']:
                    # 분야/타겟앱/AI모델은 중간 크기 (최소 15, 최대 25)
                    width = min(max(min_width, max_data_length, 15), 25)
                elif column in ['perf', 'power', 'size', 'price']:
                    # 기술 스펙은 작게 (최소 10, 최대 15)
                    width = min(max(min_width, max_data_length, 10), 15)
                elif column in ['id']:
                    # ID는 작게
                    width = 8
                else:
                    # 기본 너비
                    width = min(max(min_width, max_data_length, 15), 25)
            
            # 컬럼 너비 설정
            worksheet.column_dimensions[col_letter].width = width
            
            # created_at, updated_at 컬럼 숨기기 (이미 제외되었지만 안전장치)
            if column in ['created_at', 'updated_at']:
                worksheet.column_dimensions[col_letter].hidden = True
            
            # 텍스트 컬럼에 자동 줄바꿈 설정
            if sheet_type == 'voc' and column in ['content', 'action_item', 'ai_summary']:
                for row in range(2, worksheet.max_row + 1):  # 헤더 제외하고 데이터 행부터
                    cell = worksheet[f"{col_letter}{row}"]
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            elif sheet_type == 'project' and column in ['requirements', 'competitors', 'result', 'root_cause']:
                for row in range(2, worksheet.max_row + 1):  # 헤더 제외하고 데이터 행부터
                    cell = worksheet[f"{col_letter}{row}"]
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # 헤더 스타일 적용
        for col_num, column in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_num)
            header_cell = worksheet[f"{col_letter}1"]
            header_cell.font = header_font
            header_cell.fill = header_fill
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 행 높이 조정 (자동 줄바꿈을 위해)
        for row in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row].height = 30  # 기본 높이 설정
        
        logger.info(f"Adjusted template column widths for {sheet_type} sheet '{sheet_name}'")
        
    except Exception as e:
        logger.warning(f"Failed to adjust template column widths for sheet '{sheet_name}': {str(e)}")

def export_table_to_excel(db: Session, table_name: str, filename: Optional[str] = None) -> str:
    """
    특정 테이블만 엑셀로 내보내기

    Args:
        db: 데이터베이스 세션
        table_name: 테이블명 (users, companies, contacts, projects, vocs, audit_logs)
        filename: 파일명 (선택사항)

    Returns:
        str: 생성된 엑셀 파일의 전체 경로
    """
    try:
        # 테이블명에 따른 모델 클래스 매핑
        table_mapping = {
            'users': User,
            'companies': Company,
            'contacts': Contact,
            'projects': Project,
            'vocs': VOC,
            'audit_logs': AuditLog
        }

        if table_name not in table_mapping:
            raise ValueError(f"Unknown table name: {table_name}")

        model_class = table_mapping[table_name]

        # Export 디렉토리 확인 및 생성
        export_dir = ensure_export_directory()

        # 파일명 생성 (25XXXX 형식)
        if filename is None:
            timestamp = datetime.now().strftime("%y%m%d_%H%M")
            filename = f"export_{table_name}_{timestamp}.xlsx"

        # 전체 파일 경로 생성
        filepath = os.path.join(export_dir, filename)

        # ExcelWriter 생성
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            _export_table_to_sheet(db, model_class, table_name.title(), writer)

        logger.info(f"Excel export completed for table {table_name}: {filepath}")
        return filepath

    except Exception as e:
        logger.error(f"Excel export failed for table {table_name}: {str(e)}")
        raise

def get_table_info(db: Session) -> Dict[str, Any]:
    """
    모든 테이블의 기본 정보 반환 (레코드 수 등)
    
    Args:
        db: 데이터베이스 세션
    
    Returns:
        Dict: 테이블별 정보
    """
    try:
        table_mapping = {
            'users': User,
            'companies': Company,
            'contacts': Contact,
            'projects': Project,
            'vocs': VOC,
            'audit_logs': AuditLog
        }
        
        info = {}
        for table_name, model_class in table_mapping.items():
            count = db.query(model_class).count()
            info[table_name] = {
                'count': count,
                'columns': [column.name for column in model_class.__table__.columns]
            }
        
        return info
        
    except Exception as e:
        logger.error(f"Failed to get table info: {str(e)}")
        raise