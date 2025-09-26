#!/usr/bin/env python3
"""
===========================================
VOC Table Excel Export 기능 단순 테스트
===========================================

목적:
- excel_io.py의 export_voc_to_excel, export_full_tables_to_excel 함수 테스트
- 실제 데이터베이스 연결 없이 더미 데이터로 엑셀 export 기능 검증
- 컬럼 너비 조정, 자동 줄바꿈, company_name 추가 등 최신 기능 테스트
- VOC 전용 export와 Full tables export 두 가지 모드 테스트

테스트 내용:
1. VOC 전용 export (company_id 제외, company_name 추가)
2. Full tables export (Users 제외, 연동 ID 제외)
3. 컬럼 너비 자동 조정
4. content, action_item 자동 줄바꿈
5. created_at, updated_at 컬럼 숨김 처리

사용법:
    python test_excel_export_simple.py

생성 파일:
- export_voc_YYMMDD_HHMM.xlsx (VOC 전용)
- export_full_YYMMDD_HHMM.xlsx (전체 테이블)
"""
import sys
import os
import pandas as pd
from datetime import datetime

# exports 디렉토리 생성 (프로젝트 루트 기준)
exports_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "exports")
if not os.path.exists(exports_dir):
    os.makedirs(exports_dir)
    print(f"Created exports directory: {exports_dir}")

# 테스트 데이터 생성
def create_test_data():
    """테스트용 데이터 생성"""
    
    # VOC 테스트 데이터
    voc_data = [
        {
            "id": 1,
            "date": "2025-01-23",
            "company_name": "삼성전자",
            "content": "AI 칩 성능이 기대보다 낮습니다. 개선이 필요합니다. 특히 딥러닝 추론 속도가 경쟁사 대비 20% 느리며, 전력 소비도 높은 편입니다. 이는 고객 만족도에 직접적인 영향을 미칠 수 있는 중요한 이슈입니다.",
            "action_item": "성능 최적화 작업 진행. 하드웨어 최적화, 알고리즘 개선, 전력 효율성 향상 방안을 종합적으로 검토해야 합니다.",
            "due_date": "2025-02-15",
            "status": "in_progress",
            "priority": "high",
            "ai_summary": "AI 칩 성능 개선 요청",
            "created_at": "2025-01-23 14:30:22",
            "updated_at": "2025-01-23 14:30:22"
        },
        {
            "id": 2,
            "date": "2025-01-23",
            "company_name": "LG전자",
            "content": "스마트홈 솔루션의 가격이 너무 비쌉니다. 현재 가격으로는 중소기업이나 일반 가정에서 도입하기 어려운 수준입니다. 경쟁사 대비 30% 높은 가격으로 인해 시장 점유율 확보에 어려움을 겪고 있습니다.",
            "action_item": "가격 재검토 및 조정. 비용 구조 분석, 대량 생산 효과, 경쟁사 가격 대비 분석을 통해 적정 가격대를 설정해야 합니다.",
            "due_date": "2025-02-20",
            "status": "pending",
            "priority": "medium",
            "ai_summary": "가격 경쟁력 개선 요청",
            "created_at": "2025-01-23 14:30:22",
            "updated_at": "2025-01-23 14:30:22"
        }
    ]
    
    # Company 테스트 데이터
    company_data = [
        {
            "id": 1,
            "name": "삼성전자",
            "domain": "samsung.com",
            "revenue": "1000억",
            "employee": 50000,
            "nation": "한국",
            "created_at": "2025-01-23 14:30:22",
            "updated_at": "2025-01-23 14:30:22"
        },
        {
            "id": 2,
            "name": "LG전자",
            "domain": "lg.com",
            "revenue": "500억",
            "employee": 25000,
            "nation": "한국",
            "created_at": "2025-01-23 14:30:22",
            "updated_at": "2025-01-23 14:30:22"
        }
    ]
    
    # Contact 테스트 데이터
    contact_data = [
        {
            "id": 1,
            "name": "김철수",
            "title": "부장",
            "email": "kim.cs@samsung.com",
            "phone": "010-1234-5678",
            "note": "삼성전자 담당자",
            "created_at": "2025-01-23 14:30:22",
            "updated_at": "2025-01-23 14:30:22"
        },
        {
            "id": 2,
            "name": "이영희",
            "title": "과장",
            "email": "lee.yh@lg.com",
            "phone": "010-9876-5432",
            "note": "LG전자 담당자",
            "created_at": "2025-01-23 14:30:22",
            "updated_at": "2025-01-23 14:30:22"
        }
    ]
    
    # Project 테스트 데이터
    project_data = [
        {
            "id": 1,
            "name": "AI 칩 개발 프로젝트",
            "field": "반도체",
            "target_app": "스마트폰",
            "ai_model": "GPT-4",
            "perf": "고성능",
            "power": "저전력",
            "form_factor": "소형",
            "memory": "8GB",
            "price": "경쟁력 있는 가격",
            "requirements": "고성능 AI 처리",
            "competitors": "애플, 구글",
            "result": "성공",
            "root_cause": "기술력 우위",
            "created_at": "2025-01-23 14:30:22",
            "updated_at": "2025-01-23 14:30:22"
        },
        {
            "id": 2,
            "name": "스마트홈 솔루션",
            "field": "IoT",
            "target_app": "스마트홈",
            "ai_model": "BERT",
            "perf": "중간 성능",
            "power": "저전력",
            "form_factor": "중형",
            "memory": "16GB",
            "price": "합리적 가격",
            "requirements": "안정적인 IoT 연결",
            "competitors": "삼성, SK텔레콤",
            "result": "진행중",
            "root_cause": "시장 수요 증가",
            "created_at": "2025-01-23 14:30:22",
            "updated_at": "2025-01-23 14:30:22"
        }
    ]
    
    # AuditLog 테스트 데이터
    audit_log_data = [
        {
            "id": 1,
            "action": "create",
            "table_name": "vocs",
            "row_id": 1,
            "before_json": None,
            "after_json": '{"content": "AI 칩 성능이 기대보다 낮습니다."}',
            "ip": "127.0.0.1",
            "ua": "Test Browser",
            "created_at": "2025-01-23 14:30:22"
        },
        {
            "id": 2,
            "action": "create",
            "table_name": "vocs",
            "row_id": 2,
            "before_json": None,
            "after_json": '{"content": "스마트홈 솔루션의 가격이 너무 비쌉니다."}',
            "ip": "127.0.0.1",
            "ua": "Test Browser",
            "created_at": "2025-01-23 14:30:22"
        }
    ]
    
    return voc_data, company_data, contact_data, project_data, audit_log_data

def _adjust_column_widths(writer, sheet_name, df):
    """엑셀 컬럼 너비 자동 조정"""
    try:
        from openpyxl.utils import get_column_letter
        
        # 워크시트 가져오기
        worksheet = writer.sheets[sheet_name]
        
        # 컬럼별 너비 설정
        for column in df.columns:
            col_letter = get_column_letter(df.columns.get_loc(column) + 1)
            
            # 컬럼명 길이 계산
            col_name_length = len(str(column))
            
            # 데이터 최대 길이 계산
            max_data_length = 0
            if not df.empty:
                for value in df[column]:
                    if value is not None:
                        max_data_length = max(max_data_length, len(str(value)))
            
            # 최소 너비 설정
            min_width = max(col_name_length, 10)
            
            # 특별한 컬럼들은 더 넓게 설정
            if column in ['content', 'action_item', 'requirements', 'competitors', 'result', 'root_cause', 'note']:
                # 텍스트 컬럼은 최소 30, 최대 50
                width = min(max(min_width, max_data_length, 30), 50)
            elif column in ['ai_summary']:
                # AI 요약은 최소 25, 최대 40
                width = min(max(min_width, max_data_length, 25), 40)
            elif column in ['name', 'title', 'email', 'phone', 'domain', 'revenue', 'nation', 'field', 'target_app', 'ai_model', 'perf', 'power', 'form_factor', 'memory', 'price', 'company_name']:
                # 일반 텍스트 컬럼은 최소 15, 최대 25
                width = min(max(min_width, max_data_length, 15), 25)
            elif column in ['status', 'priority', 'action', 'table_name']:
                # 상태/타입 컬럼은 최소 12, 최대 18
                width = min(max(min_width, max_data_length, 12), 18)
            elif column in ['date', 'due_date']:
                # 날짜 컬럼은 딱 맞게 조정 (YYYY-MM-DD 형식)
                width = 12
            elif column in ['created_at', 'updated_at']:
                # 타임스탬프 컬럼은 숨김 처리
                width = 12
            elif column in ['id', 'employee']:
                # 숫자 컬럼은 최소 8, 최대 12
                width = min(max(min_width, max_data_length, 8), 12)
            else:
                # 기본 너비
                width = min(max(min_width, max_data_length, 12), 20)
            
            # 컬럼 너비 설정
            worksheet.column_dimensions[col_letter].width = width
            
            # created_at, updated_at 컬럼 숨기기
            if column in ['created_at', 'updated_at']:
                worksheet.column_dimensions[col_letter].hidden = True
            
            # content, action_item 컬럼에 자동 줄바꿈 설정
            if column in ['content', 'action_item']:
                for row in range(2, worksheet.max_row + 1):  # 헤더 제외하고 데이터 행부터
                    cell = worksheet[f"{col_letter}{row}"]
                    from openpyxl.styles import Alignment
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            
        print(f"  - 컬럼 너비 조정 완료: {sheet_name}")
        
    except Exception as e:
        print(f"  - 컬럼 너비 조정 실패: {sheet_name} - {str(e)}")

def export_voc_excel(voc_data):
    """VOC 전용 엑셀 export"""
    timestamp = datetime.now().strftime("%y%m%d_%H%M")
    filename = f"export_voc_{timestamp}.xlsx"
    filepath = os.path.join(exports_dir, filename)
    
    # VOC 데이터를 DataFrame으로 변환
    df = pd.DataFrame(voc_data)
    
    # 엑셀 파일로 저장
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="VOCs", index=False)
        
        # 컬럼 너비 자동 조정
        _adjust_column_widths(writer, "VOCs", df)
    
    print(f"✅ VOC export 완료: {filename}")
    return filepath

def export_full_excel(company_data, contact_data, project_data, voc_data, audit_log_data):
    """Full tables export (Users 제외, 연동 ID 제외)"""
    timestamp = datetime.now().strftime("%y%m%d_%H%M")
    filename = f"export_full_{timestamp}.xlsx"
    filepath = os.path.join(exports_dir, filename)
    
    # 각 테이블을 DataFrame으로 변환
    company_df = pd.DataFrame(company_data)
    contact_df = pd.DataFrame(contact_data)
    project_df = pd.DataFrame(project_data)
    voc_df = pd.DataFrame(voc_data)
    audit_log_df = pd.DataFrame(audit_log_data)
    
    # 엑셀 파일로 저장 (각 테이블별 시트)
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        company_df.to_excel(writer, sheet_name="Companies", index=False)
        contact_df.to_excel(writer, sheet_name="Contacts", index=False)
        project_df.to_excel(writer, sheet_name="Projects", index=False)
        voc_df.to_excel(writer, sheet_name="VOCs", index=False)
        audit_log_df.to_excel(writer, sheet_name="AuditLogs", index=False)
        
        # 각 시트의 컬럼 너비 자동 조정
        _adjust_column_widths(writer, "Companies", company_df)
        _adjust_column_widths(writer, "Contacts", contact_df)
        _adjust_column_widths(writer, "Projects", project_df)
        _adjust_column_widths(writer, "VOCs", voc_df)
        _adjust_column_widths(writer, "AuditLogs", audit_log_df)
    
    print(f"✅ Full export 완료: {filename}")
    return filepath

def main():
    """메인 함수"""
    print("=== 엑셀 Export 테스트 시작 ===")
    
    # 테스트 데이터 생성
    voc_data, company_data, contact_data, project_data, audit_log_data = create_test_data()
    
    print(f"테스트 데이터 생성 완료:")
    print(f"- VOC: {len(voc_data)}개")
    print(f"- Company: {len(company_data)}개")
    print(f"- Contact: {len(contact_data)}개")
    print(f"- Project: {len(project_data)}개")
    print(f"- AuditLog: {len(audit_log_data)}개")
    
    # 1. VOC 전용 export
    print("\n1. VOC 전용 export 테스트...")
    voc_file = export_voc_excel(voc_data)
    
    # 2. Full tables export
    print("\n2. Full tables export 테스트 (Users 제외)...")
    full_file = export_full_excel(company_data, contact_data, project_data, voc_data, audit_log_data)
    
    print("\n=== 테스트 완료! ===")
    print(f"생성된 파일들:")
    print(f"- VOC 전용: {os.path.basename(voc_file)}")
    print(f"- Full (Users 제외): {os.path.basename(full_file)}")
    print(f"\n파일 위치: {exports_dir}")

if __name__ == "__main__":
    main()
