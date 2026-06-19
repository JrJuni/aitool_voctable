"""Excel sanitizer 단위/통합 테스트 (수식 인젝션 방어 + 안전 파일명)."""
import os
import sys

import openpyxl
import pandas as pd

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from backend.app.excel_io import _safe_filename, _sanitize_cell, _sanitize_df


def test_sanitize_cell_escapes_formula_prefixes():
    assert _sanitize_cell("=1+1") == "'=1+1"
    assert _sanitize_cell("+cmd") == "'+cmd"
    assert _sanitize_cell("-2+3") == "'-2+3"
    assert _sanitize_cell("@SUM(A1)") == "'@SUM(A1)"
    # 선행 공백 뒤 수식 문자도 방어
    assert _sanitize_cell("   =evil()") == "'   =evil()"


def test_sanitize_cell_leaves_safe_values():
    assert _sanitize_cell("hello") == "hello"
    assert _sanitize_cell("a=b") == "a=b"  # 중간의 = 는 안전
    assert _sanitize_cell(123) == 123
    assert _sanitize_cell(None) is None


def test_sanitize_df_roundtrip_through_xlsx(tmp_path):
    df = pd.DataFrame({"name": ["=DANGER()", "safe"], "n": [1, 2]})
    path = tmp_path / "t.xlsx"
    _sanitize_df(df).to_excel(path, index=False)

    ws = openpyxl.load_workbook(path).active
    # 헤더가 1행, 데이터는 2행부터. A열 = name
    assert ws["A2"].value == "'=DANGER()"
    assert ws["A3"].value == "safe"
    assert ws["B2"].value == 1


def test_safe_filename_blocks_path_traversal():
    for evil in ("../../etc/passwd", "..\\..\\windows\\system32", "/abs/path/x"):
        out = _safe_filename(evil)
        assert "/" not in out and "\\" not in out
        assert ".." not in out
        assert out.endswith(".xlsx")


def test_safe_filename_keeps_normal_names_and_defaults():
    assert _safe_filename("export_voc_250619.xlsx") == "export_voc_250619.xlsx"
    assert _safe_filename(None) == "export.xlsx"
    assert _safe_filename("report").endswith(".xlsx")
